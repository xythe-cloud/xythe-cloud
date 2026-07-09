"""
Upload API — receives files, processes them, stores in vector DB.
"""
from fastapi import APIRouter, UploadFile, File, Form, HTTPException
from typing import Optional
import os
import shutil
from datetime import datetime
from src.services.vector_store import VectorStoreManager
from src.utils.logging import logger
import pypdf
import openpyxl
import csv
import io
from PIL import Image
import pytesseract

router = APIRouter(prefix="/api/workspace", tags=["upload"])

# Allowed file types
ALLOWED_TYPES = {
    "pdf": "application/pdf",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "xls": "application/vnd.ms-excel",
    "csv": "text/csv",
    "png": "image/png",
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
}

UPLOAD_DIR = "./data/uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)

vector_store = VectorStoreManager()

# In-memory document store (replace with database later)
document_store = {}


@router.post("/{tenant_id}/upload")
async def upload_file(
    tenant_id: str,
    file: UploadFile = File(...),
    category: str = Form("brochure")
):
    """Upload a document for processing."""
    
    # Validate file type
    ext = file.filename.split(".")[-1].lower() if "." in file.filename else ""
    if ext not in ALLOWED_TYPES:
        raise HTTPException(400, f"Unsupported file type: .{ext}")
    
    # Save file
    tenant_dir = os.path.join(UPLOAD_DIR, tenant_id)
    os.makedirs(tenant_dir, exist_ok=True)
    
    safe_filename = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{file.filename}"
    file_path = os.path.join(tenant_dir, safe_filename)
    
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)
    
    logger.info(f"File saved: {file_path} ({len(content)} bytes)")
    
    # Extract text
    try:
        text = extract_text(file_path, ext)
    except Exception as e:
        logger.error(f"Text extraction failed: {e}")
        os.remove(file_path)
        raise HTTPException(500, f"Could not extract text: {str(e)}")
    
    if not text or len(text.strip()) < 50:
        os.remove(file_path)
        raise HTTPException(400, "Document contains too little text to process")
    
    # Chunk text
    chunks = chunk_text(text)
    logger.info(f"Extracted {len(chunks)} chunks from {file.filename}")
    
    # Store in vector database
    metadata_list = [{
        "filename": file.filename,
        "category": category,
        "tenant_id": tenant_id,
        "uploaded_at": datetime.now().isoformat()
    } for _ in chunks]
    
    try:
        doc_id = len(document_store.get(tenant_id, {}).get("documents", [])) + 1
        await vector_store.add_documents(
            tenant_id=tenant_id,
            chunks=chunks,
            metadata_list=metadata_list,
            document_id=doc_id
        )
    except Exception as e:
        logger.error(f"Vector store error: {e}")
        raise HTTPException(500, f"Processing failed: {str(e)}")
    
    # Store document metadata
    if tenant_id not in document_store:
        document_store[tenant_id] = {"documents": []}
    
    doc_info = {
        "id": doc_id,
        "name": file.filename,
        "size": f"{len(content) / 1024 / 1024:.1f} MB",
        "category": category,
        "chunks": len(chunks),
        "uploaded_at": datetime.now().isoformat()
    }
    document_store[tenant_id]["documents"].append(doc_info)
    
    return {
        "status": "ok",
        "document": doc_info,
        "chunks_processed": len(chunks)
    }


@router.get("/{tenant_id}/documents")
async def get_documents(tenant_id: str):
    """Get list of uploaded documents."""
    docs = document_store.get(tenant_id, {}).get("documents", [])
    return {"documents": docs}


def extract_text(file_path: str, ext: str) -> str:
    """Extract text from various file types."""
    
    if ext == "pdf":
        text_parts = []
        reader = pypdf.PdfReader(file_path)
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
        return "\n\n".join(text_parts)
    
    elif ext in ["xlsx", "xls"]:
        text_parts = []
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            text_parts.append(f"=== {sheet_name} ===")
            for row in sheet.iter_rows(values_only=True):
                row_text = " | ".join([str(c) if c is not None else "" for c in row])
                if row_text.strip():
                    text_parts.append(row_text)
        return "\n".join(text_parts)
    
    elif ext == "csv":
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()
    
    elif ext in ["png", "jpg", "jpeg"]:
        image = Image.open(file_path)
        return pytesseract.image_to_string(image, lang="eng+msa")
    
    else:
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            return f.read()


def chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> list:
    """Split text into overlapping chunks."""
    chunks = []
    paragraphs = text.split("\n\n")
    
    current_chunk = ""
    for para in paragraphs:
        para = para.strip()
        if not para:
            continue
        
        if len(current_chunk) + len(para) > chunk_size and current_chunk:
            chunks.append(current_chunk.strip())
            current_chunk = current_chunk[-overlap:] if len(current_chunk) > overlap else ""
        
        current_chunk += para + "\n\n"
    
    if current_chunk.strip():
        chunks.append(current_chunk.strip())
    
    return chunks